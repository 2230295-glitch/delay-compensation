"""
지체보상금 대시보드 v6
실행: streamlit run app.py
"""
import io, re, calendar
from datetime import date
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as gcl

st.set_page_config(page_title="지체보상금 대시보드", layout="wide", page_icon="📊")

# 나눔고딕 <link> 태그로 로드 (더 안정적)
st.markdown(
    '<link rel="preconnect" href="https://fonts.googleapis.com">'
    '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
    '<link href="https://fonts.googleapis.com/css2?family=Nanum+Gothic:wght@400;700;800&display=swap" rel="stylesheet">',
    unsafe_allow_html=True,
)

st.markdown("""<style>
html, body, [class*="css"], .stApp, button, input, select, textarea,
[data-testid], [data-baseweb], .stMarkdown, .stDataFrame, p, span, div, th, td {
  font-family: 'Nanum Gothic', 'Malgun Gothic', sans-serif !important;
}

/* ── 배경 + 상단 패딩 제거 ── */
.stApp { background: #f2f4f8; }
header[data-testid="stHeader"] { display: none !important; }
#MainMenu { display: none !important; }
.stMainBlockContainer, .block-container, .appview-container .main .block-container {
  max-width: 100% !important;
  padding: 0 2rem 3rem !important;
  padding-top: 0 !important;
}

/* ── 사이드바 ── */
section[data-testid="stSidebar"] > div:first-child { background: #1b2838 !important; }
section[data-testid="stSidebar"] * {
  font-family: 'Nanum Gothic', 'Malgun Gothic', sans-serif !important;
}
section[data-testid="stSidebar"] label { color: #8ab3cf !important; font-size: .82rem !important; }
section[data-testid="stSidebar"] p     { color: #6a93b0 !important; font-size: .78rem !important; }
section[data-testid="stSidebar"] hr    { border-color: #243444 !important; }
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
  background: #162432 !important; border: 1.5px dashed #2e4f6a !important; border-radius: 8px !important;
}

/* ── 헤더 ── */
.ph {
  background: #1b2838;
  padding: 1rem 2rem; margin: 0 -2rem 1.6rem;
  display: flex; align-items: center; justify-content: space-between;
  border-bottom: 3px solid #e67e22;
}
.ph-l { display: flex; flex-direction: column; gap: .15rem; }
.ph-title { font-size: 1.3rem; font-weight: 800; color: #fff; }
.ph-sub   { font-size: .72rem; color: #6a9dbf; }
.ph-badge {
  background: #243444; color: #7ab8d8; font-size: .72rem; font-weight: 700;
  padding: .3rem .85rem; border-radius: 4px; border: 1px solid #2e4a62;
}

/* ── KPI 카드 4개 ── */
.kpi-wrap {
  display: grid; grid-template-columns: repeat(4, 1fr);
  gap: 1rem; margin-bottom: 1.4rem;
}
.kpi-card {
  background: #fff; border-radius: 10px;
  padding: 1.15rem 1.4rem 1rem;
  box-shadow: 0 1px 6px rgba(0,0,0,.07);
  border-top: 4px solid #ccc;
  position: relative;
}
.kpi-card.c-orange { border-top-color: #e67e22; }
.kpi-card.c-red    { border-top-color: #c0392b; }
.kpi-card.c-blue   { border-top-color: #2471a3; }
.kpi-card.c-gray   { border-top-color: #5d6d7e; }

.kpi-lbl  { font-size: .72rem; color: #999; font-weight: 700; letter-spacing: .04em; margin-bottom: .4rem; }
.kpi-num  { font-size: 1.95rem; font-weight: 800; line-height: 1.1; color: #1a2535; }
.kpi-unit { font-size: 1rem; font-weight: 700; color: #555; margin-left: .1rem; }
.kpi-delta { font-size: .76rem; margin-top: .35rem; color: #aaa; }
.kpi-up   { color: #c0392b; font-weight: 700; }
.kpi-dn   { color: #2471a3; font-weight: 700; }

/* ── 섹션 제목 ── */
.sec-hd {
  font-size: .9rem; font-weight: 800; color: #1a2535;
  display: flex; align-items: center; gap: .45rem;
  padding-bottom: .5rem; margin-bottom: .85rem;
  border-bottom: 1.5px solid #e5e9f0;
}
.sec-bar { width: 4px; height: 16px; border-radius: 2px; background: #2471a3; flex-shrink: 0; }

/* ── 다운로드 버튼 ── */
.stDownloadButton > button {
  background: #1b2838 !important; color: #fff !important;
  border: none !important; border-radius: 6px !important;
  font-weight: 800 !important; font-size: .84rem !important;
  font-family: 'Nanum Gothic', 'Malgun Gothic', sans-serif !important;
  padding: .55rem 1.1rem !important;
}
.stDownloadButton > button:hover { background: #243444 !important; }

/* ── 탭 ── */
.stTabs [data-baseweb="tab-list"] { gap: .25rem; border-bottom: 2px solid #dde3ef; }
.stTabs [data-baseweb="tab"] {
  font-weight: 700; font-size: .84rem; padding: .38rem 1rem; color: #aaa;
  font-family: 'Nanum Gothic', 'Malgun Gothic', sans-serif !important;
}
.stTabs [aria-selected="true"] { color: #1b2838 !important; border-bottom: 2px solid #1b2838 !important; }

/* ── 버튼 ── */
.stButton > button {
  font-family: 'Nanum Gothic', 'Malgun Gothic', sans-serif !important;
  font-weight: 700 !important; border-radius: 6px !important;
}

/* ── border 컨테이너 ── */
[data-testid="stVerticalBlockBorderWrapper"] {
  background: #fff !important;
  border-radius: 10px !important;
  border: 1px solid #e0e5ef !important;
  box-shadow: 0 1px 6px rgba(0,0,0,.06) !important;
}

/* ── 연도/월 선택기 ── */
.yr-mo-wrap { display:flex; align-items:flex-start; gap:1.2rem; margin-bottom:1.2rem; }
.yr-btns { display:flex; flex-direction:column; gap:.4rem; }
.mo-grid { display:grid; grid-template-columns:repeat(6,1fr); gap:.35rem; flex:1; }
.mo-btn-outer button, .yr-btn-outer button {
  border-radius:6px !important; font-weight:700 !important;
  font-size:.82rem !important; padding:.35rem .5rem !important;
  transition: all .15s !important;
}
</style>""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
ALLOW_CH       = {"DW04","DW07","DW08","DW09","DW10","DW11","DW17"}
EXCLUDE_CODES  = {"10014461","C4900","C2000","C3100"}
EXCLUDE_BU     = {"기타사업부","수탁사업부","완제사업부"}
EXCLUDE_OFFICE = {"DW_B2B","DW_CH신사업팀"}
EXCLUDE_MGR    = {"1202150261"}
FONT           = "Nanum Gothic,Malgun Gothic,sans-serif"

def _num(v):
    try: return float(str(v).replace(",","").strip())
    except: return float("nan")

def fmt_won(v, short=False):
    v = int(v)
    if short:
        if abs(v) >= 100_000_000: return f"{v/100_000_000:.1f}억"
        if abs(v) >= 10_000:      return f"{round(v/10_000):,}만"
        return f"{v:,}"
    if abs(v) >= 100_000_000: return f"{v/100_000_000:.2f}억원"
    if abs(v) >= 10_000:      return f"{round(v/10_000):,}만원"
    return f"{v:,}원"

def mo_label(s):
    y, m = s.split("-")
    return f"{y[2:]}년 {int(m)}월"

def _norm(s):
    s = re.sub(r'[\s\(\)\.\-·\']', '', str(s)).upper()
    for p in ['의료법인','재단법인','학교법인','사회복지법인']:
        if s.startswith(p): s = s[len(p):]
    s = re.sub(r'^[의재갑을](?=[가-힣])', '', s)
    return s

# ──────────────────────────────────────────────────────────────
def load_contract(file_bytes, sap_df):
    try:    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='계약서', dtype=str, header=5)
    except: df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, header=5)
    cols = ['No','병원명','계약유형','사업자번호','계약일자','계약병상','계약금액','계약기간',
            '_8','대금지급','지체보상금조항','비고','_12']
    df.columns = cols + [f'_e{i}' for i in range(max(0, len(df.columns)-len(cols)))]
    df = df[df['No'].notna() & df['No'].astype(str).str.match(r'^\d+$', na=False)].copy()
    df['지체여부'] = df['지체보상금조항'].apply(lambda x: '지체보상금' in str(x))
    def _rate(x):
        m = re.search(r'\[(\d+)/(\d+)\]', str(x))
        return float(m.group(1)) / float(m.group(2)) if m else 1/1000
    df['요율'] = df['지체보상금조항'].apply(_rate)
    df['_norm'] = df['병원명'].apply(_norm)
    s2 = sap_df.copy(); s2['_norm'] = s2['판매처명'].apply(_norm)
    sm = dict(zip(s2['_norm'], s2['판매처']))
    df['코드'] = df['_norm'].map(sm)
    for idx, row in df[df['코드'].isna()].iterrows():
        cn = row['_norm']
        if len(cn) < 4: continue
        for sn, sc in zip(s2['_norm'], s2['판매처']):
            if len(sn) >= 4 and (cn in sn or sn in cn):
                df.at[idx, '코드'] = sc; break
    result = {}
    for _, r in df[df['코드'].notna()].iterrows():
        c = r['코드']
        if c not in result or not r['지체여부']:
            result[c] = {"지체여부": r['지체여부'], "요율": r['요율']}
    n_yes = sum(1 for v in result.values() if v["지체여부"])
    n_no  = sum(1 for v in result.values() if not v["지체여부"])
    return result, n_yes, n_no

def load_base(file_bytes, cmap=None, think_rotations=None):
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str).dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    out = []
    for _, r in df.iterrows():
        code = str(r.get("판매처","")).strip(); name = str(r.get("판매처명","")).strip()
        base = _num(r.get("기준회전일",0)); yak = _num(r.get("약정회전일",0))
        exc  = _num(r.get("예외회전일",0))
        if not code or code == "nan" or pd.isna(base) or base <= 0: continue
        eff = (int(exc) if not pd.isna(exc) and exc > 0
               else (int(yak) if not pd.isna(yak) and yak > 0 else int(base)))
        info = (cmap or {}).get(code, {})
        use = info.get("지체여부", True); rate = info.get("요율", 1/1000)
        out.append({"코드":code,"명칭":name,"기준회전일":int(base),
                    "유효기준회전일":eff,"지체여부":use,"요율":rate})
    existing = {r["코드"] for r in out}
    if think_rotations:
        for code, rot in think_rotations.items():
            if code not in existing:
                info = (cmap or {}).get(code, {})
                use = info.get("지체여부", True); rate = info.get("요율", 1/1000)
                out.append({"코드":code,"명칭":"","기준회전일":rot,
                            "유효기준회전일":rot,"지체여부":use,"요율":rate})
    return pd.DataFrame(out)

def _parse_rot(v):
    s = str(v).strip()
    if '개월' in s:
        try: return int(s.replace('개월','').strip()) * 30
        except: return 30
    if '일' in s:
        try: return int(s.replace('일','').strip())
        except: return 30
    try: return int(s)
    except: return 30

def load_think(file_bytes):
    """씽크현황에서 씽크 거래처 코드 + 기준회전일 반환 (load_base 보완용)"""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='씽크현황', header=None, dtype=str)
    codes = {}
    for _, r in df.iloc[4:].iterrows():
        code = str(r.iloc[2]).strip()
        if not code or code in ('nan', '판매처', 'None'): continue
        codes[code] = _parse_rot(r.iloc[8])
    return codes

def load_think_monthly(file_bytes):
    """씽크현황 월별 매출/수금 → 누적잔고·현회전일 역산, calculate()용 months 반환"""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='씽크현황', header=None, dtype=str)
    row2 = df.iloc[2].tolist()
    row3 = df.iloc[3].tolist()

    # 월별 컬럼 인덱스 수집 {mo_str: {'매출': col_idx, '수금': col_idx}}
    mo_cols = {}
    cur = None
    for i, (r2, r3) in enumerate(zip(row2, row3)):
        s2 = str(r2).strip()
        if '.' in s2 and s2 not in ('nan','None'):
            cur = s2; mo_cols[cur] = {}
        if cur and str(r3).strip() in ('매출','수금') and str(r3).strip() not in mo_cols[cur]:
            mo_cols[cur][str(r3).strip()] = i

    # 유효한 연월만 (합계·계약금액 제외)
    valid = {k: v for k, v in mo_cols.items()
             if '합계' not in k and '계약' not in k and len(k) == 7}

    months_dict = {}
    for _, row in df.iloc[4:].iterrows():
        code = str(row.iloc[2]).strip()
        name = str(row.iloc[3]).strip()
        if not code or code in ('nan','None','판매처'): continue
        기준 = _parse_rot(row.iloc[8])

        cum_sale = cum_coll = 0.0
        prev_현회전일 = 0
        for mo_str, cidx in sorted(valid.items()):
            try: yr, mo = int(mo_str.split('.')[0]), int(mo_str.split('.')[1])
            except: continue
            def _v(col):
                try:
                    s = str(row.iloc[col]).replace(',','').strip()
                    if s.lower() in ('nan','none',''): return 0.0
                    return float(s) * 1_000_000
                except: return 0.0
            sale = _v(cidx['매출']) if '매출' in cidx else 0.0
            coll = _v(cidx['수금']) if '수금' in cidx else 0.0
            cum_sale += sale; cum_coll += coll
            잔고 = cum_sale - cum_coll
            days = calendar.monthrange(yr, mo)[1]  # 해당 월 실제 일수
            if 잔고 <= 0:
                prev_현회전일 = 0
                continue
            if sale > 0:
                # 세금계산서 월말 발행 → 발행 당월은 회전일 1일로 시작
                현회전일 = 1
            else:
                # 수금 없으면 해당 월 실제 일수만큼 지연 누적
                현회전일 = prev_현회전일 + days
            prev_현회전일 = 현회전일
            key = (yr, mo)
            if key not in months_dict: months_dict[key] = []
            months_dict[key].append({
                "연": yr, "월": mo, "코드": code, "명칭": name,
                "채널": "씽크", "잔고": int(잔고), "현회전일": 현회전일,
                "_기준회전일_override": 기준,
                "사무소": "", "사업부": "", "담당자": "",
            })

    months = [{"yr": yr, "mo": mo, "data": pd.DataFrame(rows)}
              for (yr, mo), rows in sorted(months_dict.items())]
    return months

def load_monthly(file_bytes, think_codes=None):
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str).dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    def _ym(v):
        s = str(v).strip()
        return (int(s.split(".")[0]), int(s.split(".")[1])) if "." in s else (None, None)
    df["_yr"], df["_mo"] = zip(*df["년월"].map(_ym))
    df = df[df["_yr"].notna()].copy()
    df["_yr"] = df["_yr"].astype(int); df["_mo"] = df["_mo"].astype(int)
    df["_잔고"] = df["순실잔고"].apply(_num); df["_일수"] = df["순실회전일"].apply(_num)
    df = df.dropna(subset=["_잔고","_일수"]).copy(); df["_일수"] = df["_일수"].astype(int)
    df = df[df["판매처"].notna() & (df["판매처"].str.strip() != "") & (df["판매처"] != "nan")]
    df = df[df["채널"].isin(ALLOW_CH)]; df = df[df["_잔고"] != 0]
    if think_codes and "채널" in df.columns:
        dw07_mask = df["채널"] == "DW07"
        df = df[~dw07_mask | df["판매처"].isin(think_codes)]
    if "담당자"   in df.columns: df = df[~df["담당자"].isin(EXCLUDE_MGR)]
    if "사업부명" in df.columns: df = df[~df["사업부명"].apply(lambda x: any(k in str(x) for k in EXCLUDE_BU))]
    df = df[~df["판매처"].isin(EXCLUDE_CODES)]
    if "사무소명" in df.columns: df = df[~df["사무소명"].isin(EXCLUDE_OFFICE)]
    months = []
    for (yr, mo), grp in df.groupby(["_yr","_mo"]):
        rows = [{"연":int(yr),"월":int(mo),"코드":str(r["판매처"]).strip(),
                 "명칭":str(r.get("판매처명","")).strip(),"채널":str(r.get("채널","")).strip(),
                 "잔고":r["_잔고"],"현회전일":r["_일수"],
                 "사무소":str(r.get("사무소명","")).strip() if "사무소명" in r.index else "",
                 "사업부":str(r.get("사업부명","")).strip() if "사업부명" in r.index else "",
                 "담당자":str(r.get("담당자명","")).strip() if "담당자명" in r.index else "",
                 } for _, r in grp.iterrows()]
        months.append({"yr":int(yr),"mo":int(mo),"data":pd.DataFrame(rows)})
    months.sort(key=lambda x: (x["yr"], x["mo"]))
    return months

def calculate(months, base_df):
    bmap = {r["코드"]: r for _, r in base_df.iterrows()}
    cum: dict = {}; results = []
    for m in months:
        yr, mo = m["yr"], m["mo"]
        for _, row in m["data"].iterrows():
            code = row["코드"]
            if code not in bmap: continue
            b = bmap[code]
            if not b["지체여부"]: continue
            기준 = int(row["_기준회전일_override"]) if "_기준회전일_override" in row and pd.notna(row["_기준회전일_override"]) else b["유효기준회전일"]
            요율 = b["요율"]
            현 = row["현회전일"]; 잔 = row["잔고"]
            지연 = max(0, 현 - 기준); prev = cum.get(code, 0)
            if 지연 <= 0: cum[code] = 0; 증분 = 0
            else: 증분 = max(0, 지연 - prev); cum[code] = max(prev, 지연)
            보상 = round(잔 * 증분 * 요율) if 증분 > 0 else 0
            rate_str = f"1/{int(round(1/요율))}" if 요율 > 0 else "-"
            results.append({
                "기준월":f"{yr}-{mo:02d}","판매처코드":code,"판매처명":row["명칭"],
                "채널":row["채널"],
                "사무소":row.get("사무소",""),"사업부":row.get("사업부",""),"담당자":row.get("담당자",""),
                "현 미수금":int(잔),"기준회전일(SAP)":b["기준회전일"],
                "유효기준회전일":기준,"현 회전일":현,"지연일수(누적)":지연,
                "청구 증분일수":증분,"지체보상금":보상,"요율":rate_str,
                "비고":"청구" if 보상 > 0 else ("정상화" if 지연 == 0 and prev > 0 else ""),
            })
    return pd.DataFrame(results)

def make_excel(df, title="전체내역"):
    wb = openpyxl.Workbook(); HDR = PatternFill("solid", fgColor="18293f")
    def _h(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = HDR; c.alignment = Alignment(horizontal="center", vertical="center")
    w = {"기준월":10,"판매처코드":12,"판매처명":22,"채널":10,"현 미수금":14,
         "기준회전일(SAP)":12,"유효기준회전일":12,"현 회전일":10,"지연일수(누적)":12,
         "청구 증분일수":12,"지체보상금":14,"요율":8,"비고":8}
    ws = wb.active; ws.title = title
    cols = list(df.columns); ws.append(cols)
    for r in df.itertuples(index=False): ws.append(list(r))
    _h(ws)
    for ci, cn in enumerate(cols, 1):
        for ri in range(2, ws.max_row+1):
            c = ws.cell(ri, ci)
            if cn in {"현 미수금","지체보상금"}: c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[gcl(ci)].width = w.get(cn, 12)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def make_full_excel(df):
    wb = openpyxl.Workbook(); HDR = PatternFill("solid", fgColor="18293f")
    def _h(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = HDR; c.alignment = Alignment(horizontal="center", vertical="center")
    w = {"기준월":10,"판매처코드":12,"판매처명":22,"채널":10,"현 미수금":14,
         "기준회전일(SAP)":12,"유효기준회전일":12,"현 회전일":10,"지연일수(누적)":12,
         "청구 증분일수":12,"지체보상금":14,"요율":8,"비고":8}
    ws1 = wb.active; ws1.title = "전체내역"
    cols = list(df.columns); ws1.append(cols)
    for r in df.itertuples(index=False): ws1.append(list(r))
    _h(ws1)
    for ci, cn in enumerate(cols, 1):
        for ri in range(2, ws1.max_row+1):
            c = ws1.cell(ri, ci)
            if cn in {"현 미수금","지체보상금"}: c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[gcl(ci)].width = w.get(cn, 12)
    ws1.freeze_panes = "A2"
    ws2 = wb.create_sheet("거래처별누계")
    pv = df.pivot_table(index=["판매처코드","판매처명"], columns="기준월",
                        values="지체보상금", aggfunc="sum", fill_value=0).reset_index()
    pc = list(pv.columns); ws2.append(pc)
    for r in pv.itertuples(index=False): ws2.append(list(r))
    _h(ws2)
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, (int, float)) and c.value != 0: c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 12; ws2.column_dimensions["B"].width = 22
    for i in range(3, len(pc)+1): ws2.column_dimensions[gcl(i)].width = 14
    ws2.freeze_panes = "C2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ──────────────────────────────────────────────────────────────
BASE_PATH  = Path("input/기준회전일/SAP 거래처별 기준회전일_DW11_0722.xlsx")
MO_FOLDER  = Path("input/월별잔고회전일")
THINK_PATH = next(iter(sorted(Path("input").glob("씽크*.xlsx")) + sorted(Path("input").glob("씽크*.XLSX"))), None)
CT_LIST    = sorted(Path("input/계약정보").glob("*.xlsx")) + sorted(Path("input/계약정보").glob("*.xls"))
CT_PATH    = CT_LIST[0] if CT_LIST else None

@st.cache_data(show_spinner=False)
def _read_mo_folder(folder_str: str):
    folder = Path(folder_str)
    files = sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xls"))
    if not files:
        return None
    dfs = [pd.read_excel(f, dtype=str).dropna(how="all") for f in files]
    combined = pd.concat(dfs, ignore_index=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        combined.to_excel(w, index=False)
    return buf.getvalue(), [f.name for f in files]

@st.cache_data(show_spinner=False)
def _read_bytes(path_str: str):
    p = Path(path_str)
    return p.read_bytes() if p.exists() else None

# ──────────────────────────────────────────────────────────────
# 사이드바
# ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div style="background:#080f18;margin:-1rem -1rem .8rem;padding:1.1rem 1.2rem;">'
        '<div style="font-size:1rem;font-weight:800;color:#fff;font-family:\'Nanum Gothic\',sans-serif">📊 지체보상금</div>'
        '<div style="font-size:.68rem;color:#3a6885;margin-top:.2rem;font-family:\'Nanum Gothic\',sans-serif">DW11 씽크채널 · 누적 증분 방식</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    use_default = st.checkbox("input 폴더 기본 파일 사용", value=True)
    if use_default:
        _mo_files = (sorted(MO_FOLDER.glob("*.xlsx")) + sorted(MO_FOLDER.glob("*.xls"))) if MO_FOLDER.exists() else []
        _mo_lbl = f"월별 잔고 ({len(_mo_files)}개 파일)" if _mo_files else "월별 잔고"
        for lbl, ok in [(_mo_lbl, bool(_mo_files)),
                         ("기준회전일", BASE_PATH.exists()),
                         ("계약서", CT_PATH is not None)]:
            c = "#3dbb85" if ok else "#e05555"
            i = "✓" if ok else "✗"
            st.markdown(f'<div style="font-size:.78rem;color:{c};padding:.15rem 0;'
                        f'font-family:\'Nanum Gothic\',sans-serif">{i} {lbl}</div>',
                        unsafe_allow_html=True)
    else:
        f_monthly  = st.file_uploader("월별 잔고",   type=["xlsx","xls"])
        f_base     = st.file_uploader("기준회전일",  type=["xlsx","xls"])
        f_contract = st.file_uploader("계약서 파일", type=["xlsx","xls"])
    st.markdown("---")
    st.markdown(
        '<div style="font-size:.72rem;color:#3a5e7a;line-height:1.9;font-family:\'Nanum Gothic\',sans-serif">'
        '<b style="color:#5a8faa">계산 기준</b><br>'
        '월말 순실잔고 × 증분지연일수 × 요율<br>'
        '<b style="color:#5a8faa">유효기준회전일</b><br>'
        '예외 > 약정 > 기준</div>',
        unsafe_allow_html=True,
    )

# ──────────────────────────────────────────────────────────────
# 데이터
# ──────────────────────────────────────────────────────────────
if use_default:
    _mo_result = _read_mo_folder(str(MO_FOLDER)) if MO_FOLDER.exists() else None
    m_b, _mo_names = _mo_result if _mo_result else (None, [])
    b_b = _read_bytes(str(BASE_PATH))
    c_b = _read_bytes(str(CT_PATH)) if CT_PATH else None
    t_b = _read_bytes(str(THINK_PATH)) if THINK_PATH else None
else:
    m_b = f_monthly.read()  if f_monthly  else None
    _mo_names = [f_monthly.name] if f_monthly else []
    b_b = f_base.read()     if f_base     else None
    c_b = f_contract.read() if f_contract else None
    t_b = THINK_PATH.read_bytes() if THINK_PATH else None

def _build_org_map(m_b):
    """SAP 월별 파일에서 판매처코드 → {사무소, 사업부, 담당자} 매핑"""
    if not m_b:
        return {}
    try:
        df = pd.read_excel(io.BytesIO(m_b), dtype=str).dropna(how="all")
        df.columns = [str(c).strip() for c in df.columns]
        org_map = {}
        for _, r in df.iterrows():
            code = str(r.get("판매처","")).strip()
            if not code or code == "nan":
                continue
            org_map[code] = {
                "사무소": str(r.get("사무소명","")).strip() if "사무소명" in r.index else "",
                "사업부": str(r.get("사업부명","")).strip() if "사업부명" in r.index else "",
                "담당자": str(r.get("담당자명","")).strip() if "담당자명" in r.index else "",
            }
        return org_map
    except Exception:
        return {}

@st.cache_data(show_spinner=False)
def run_calc(m_b, b_b, c_b, t_b):
    sap_raw = pd.read_excel(io.BytesIO(b_b), dtype=str).dropna(how="all")
    sap_raw.columns = [str(c).strip() for c in sap_raw.columns]
    cmap = n_yes = n_no = None
    if c_b:
        cmap, n_yes, n_no = load_contract(c_b, sap_raw)
    think_rot = load_think(t_b) if t_b else None
    bdf = load_base(b_b, cmap, think_rot)
    months = load_think_monthly(t_b) if t_b else load_monthly(m_b)
    org_map = _build_org_map(m_b)
    result = calculate(months, bdf)
    # org 정보 결과에 병합
    if org_map and not result.empty:
        result["사무소"] = result["판매처코드"].map(lambda c: org_map.get(c, {}).get("사무소", ""))
        result["사업부"] = result["판매처코드"].map(lambda c: org_map.get(c, {}).get("사업부", ""))
        result["담당자"] = result["판매처코드"].map(lambda c: org_map.get(c, {}).get("담당자", ""))
    return result, cmap, n_yes, n_no

# ──────────────────────────────────────────────────────────────
# 헤더
# ──────────────────────────────────────────────────────────────
today_str = date.today().strftime("%Y. %m. %d")
st.markdown(f"""<div class="ph">
  <div class="ph-l">
    <div class="ph-title">지체보상금 계산 대시보드</div>
    <div class="ph-sub">DW11 씽크채널 · 월말 순실잔고 × 증분지연일수 × 요율</div>
  </div>
  <div class="ph-badge">{today_str} 기준</div>
</div>""", unsafe_allow_html=True)

result_df = contract_map = None
if m_b and b_b:
    try:
        with st.spinner("계산 중..."):
            result_df, contract_map, n_yes, n_no = run_calc(m_b, b_b, c_b, t_b)
    except Exception as e:
        st.error(f"오류: {e}"); st.stop()
else:
    st.info("왼쪽에서 파일을 선택하거나 업로드하세요."); st.stop()

if result_df is None or result_df.empty:
    st.warning("계산 결과가 없습니다."); st.stop()


months_sorted = sorted(result_df["기준월"].unique())
mo_sum = result_df.groupby("기준월")["지체보상금"].sum()

# ── 월 선택 ──────────────────────────────────────────────────
if "sel_mo" not in st.session_state or st.session_state.sel_mo not in months_sorted:
    st.session_state.sel_mo = months_sorted[-1]

# 연도 목록 추출
years_avail = sorted(set(m.split("-")[0] for m in months_sorted))
if "sel_yr" not in st.session_state or st.session_state.sel_yr not in years_avail:
    st.session_state.sel_yr = st.session_state.sel_mo.split("-")[0]

with st.container():
    st.markdown('<div style="font-size:.7rem;font-weight:800;color:#a0aec0;letter-spacing:.1em;margin-bottom:.6rem">기준월 선택</div>', unsafe_allow_html=True)
    yr_c = st.columns([1]*len(years_avail) + [6-len(years_avail)] if len(years_avail) < 6 else [1]*len(years_avail))
    for i, yr in enumerate(years_avail):
        with yr_c[i]:
            if st.button(f"{yr}년", key=f"yr_{yr}",
                         type="primary" if yr == st.session_state.sel_yr else "secondary",
                         use_container_width=True):
                st.session_state.sel_yr = yr
                yr_months = [m for m in months_sorted if m.startswith(yr)]
                if st.session_state.sel_mo not in yr_months:
                    st.session_state.sel_mo = yr_months[-1]
                st.rerun()

    yr_months = [m for m in months_sorted if m.startswith(st.session_state.sel_yr)]
    mo_cols = st.columns(6)
    for i in range(6):
        if i < len(yr_months):
            m = yr_months[i]
            mo_num = int(m.split("-")[1])
            with mo_cols[i]:
                if st.button(f"{mo_num}월", key=f"mo_{m}",
                             type="primary" if m == st.session_state.sel_mo else "secondary",
                             use_container_width=True):
                    st.session_state.sel_mo = m; st.rerun()
        else:
            mo_cols[i].empty()

sel     = st.session_state.sel_mo
sel_idx = months_sorted.index(sel)
prev    = months_sorted[sel_idx-1] if sel_idx > 0 else None
sel_df  = result_df[result_df["기준월"] == sel]
prev_df = result_df[result_df["기준월"] == prev] if prev else pd.DataFrame()

sel_total  = int(sel_df["지체보상금"].sum())
sel_n      = int((sel_df["지체보상금"] > 0).sum())
sel_cu     = sel_df[sel_df["지체보상금"] > 0]["판매처코드"].nunique()
prev_total = int(prev_df["지체보상금"].sum()) if not prev_df.empty else 0

_sel_yr = sel.split("-")[0]
ann_total = int(result_df[result_df["기준월"].str.startswith(_sel_yr)]["지체보상금"].sum())
prev_n    = int((prev_df["지체보상금"] > 0).sum()) if not prev_df.empty else 0
prev_cu   = prev_df[prev_df["지체보상금"] > 0]["판매처코드"].nunique() if not prev_df.empty else 0

def _delta(cur, prv, unit="원"):
    if prv == 0: return '<span style="color:#bbb">— 전월 비교 없음</span>'
    d = cur - prv
    if d == 0: return '<span style="color:#bbb">± 0 전월 동일</span>'
    arr = "▲" if d > 0 else "▼"
    cls = "kpi-up" if d > 0 else "kpi-dn"
    val = fmt_won(abs(d)) if unit == "원" else f"{abs(d)}{unit}"
    return f'<span class="{cls}">{arr} {val} 전월비</span>'

# 전월비 숫자 추출용
d_amt  = sel_total - prev_total
d_unit = "원"
d_arrow = "▲" if d_amt > 0 else ("▼" if d_amt < 0 else "")
d_cls   = "kpi-up" if d_amt > 0 else ("kpi-dn" if d_amt < 0 else "")
d_txt   = f'<span class="{d_cls}">{d_arrow} {fmt_won(abs(d_amt))}</span>' if d_amt != 0 and prev_total > 0 else '<span style="color:#bbb">— 전월 비교 없음</span>'

st.markdown(f"""
<div class="kpi-wrap">
  <div class="kpi-card c-orange">
    <div class="kpi-lbl">{mo_label(sel)} 지체보상금</div>
    <div class="kpi-num">{fmt_won(sel_total,short=True)}</div>
    <div class="kpi-delta">{d_txt}</div>
  </div>
  <div class="kpi-card c-red">
    <div class="kpi-lbl">청구 건수</div>
    <div class="kpi-num">{sel_n}<span class="kpi-unit">건</span></div>
    <div class="kpi-delta">{_delta(sel_n, prev_n, '건')}</div>
  </div>
  <div class="kpi-card c-blue">
    <div class="kpi-lbl">청구 거래처</div>
    <div class="kpi-num">{sel_cu}<span class="kpi-unit">개</span></div>
    <div class="kpi-delta">{_delta(sel_cu, prev_cu, '개')}</div>
  </div>
  <div class="kpi-card c-gray">
    <div class="kpi-lbl">{_sel_yr}년 누계 지체보상금</div>
    <div class="kpi-num">{fmt_won(ann_total,short=True)}</div>
    <div class="kpi-delta" style="color:#888">{_sel_yr}년 1월 ~ {mo_label(sel)}</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── 신규연체 / 추가연체 / 수금완료 ────────────────────────────
_yr2, _mo2 = int(sel.split("-")[0]), int(sel.split("-")[1])
_prev_mo = f"{_yr2-1}-12" if _mo2 == 1 else f"{_yr2}-{_mo2-1:02d}"
_prev_df = result_df[result_df["기준월"] == _prev_mo]
_prev_map = {r["판매처코드"]: r for _, r in _prev_df.iterrows()}
_curr_map = {r["판매처코드"]: r for _, r in sel_df.iterrows()}

_신규, _추가, _수금 = [], [], []
for _code in set(_prev_map) | set(_curr_map):
    _c = _curr_map.get(_code)
    _p = _prev_map.get(_code)
    _cj = int(_c["지체보상금"]) if _c is not None else 0
    _pj = int(_p["지체보상금"]) if _p is not None else 0
    _cp = int(_c["현 미수금"]) if _c is not None else 0
    _pp = int(_p["현 미수금"]) if _p is not None else 0
    _nm = (_c if _c is not None else _p)["판매처명"]
    if _cj > 0 and _pj == 0:
        _신규.append({"명칭": _nm, "금액": _cj})
    elif _cj > 0 and _pj > 0:
        _추가.append({"명칭": _nm, "금액": _cj})
    if _pj > 0 and _cj == 0:
        _수금.append({"명칭": _nm, "금액": max(_pp - _cp, 0) or _pp})

for _lst in (_신규, _추가, _수금):
    _lst.sort(key=lambda x: -x["금액"])

def _cmp_card(title, lst, accent, sign, amt_color):
    total = sum(x["금액"] for x in lst)
    rows = ""
    for i, x in enumerate(lst[:5]):
        rows += (
            f'<div style="display:flex;justify-content:space-between;align-items:center;'
            f'padding:.42rem 0;border-bottom:1px solid #f2f2f2">'
            f'<div style="display:flex;align-items:center;gap:.5rem;min-width:0;flex:1">'
            f'<span style="font-size:.68rem;color:#ccc;width:.8rem;flex-shrink:0">{i+1}</span>'
            f'<span style="font-size:.83rem;color:#2d2d2d;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{x["명칭"]}</span>'
            f'</div>'
            f'<span style="font-size:.84rem;font-weight:700;color:{amt_color};white-space:nowrap;margin-left:.5rem">'
            f'{sign}{fmt_won(x["금액"],short=True)}</span>'
            f'</div>'
        )
    if len(lst) > 5:
        etc = lst[5:]; etc_tot = sum(e["금액"] for e in etc)
        rows += f'<div style="font-size:.72rem;color:#bbb;padding:.4rem 0 0">그 외 소액 {len(etc)}처 ({fmt_won(etc_tot,short=True)})</div>'
    if not lst:
        rows = '<div style="color:#ccc;font-size:.82rem;padding:1.8rem 0;text-align:center">해당 없음</div>'
    return (
        f'<div style="background:#fff;border-radius:10px;border-top:3px solid {accent};'
        f'box-shadow:0 1px 8px rgba(0,0,0,.07);padding:1rem 1.15rem 1.1rem">'
        f'<div style="display:flex;align-items:baseline;gap:.5rem;margin-bottom:.65rem">'
        f'<span style="font-size:.88rem;font-weight:800;color:{accent}">{title}</span>'
        f'<span style="font-size:.76rem;color:#aaa">{len(lst)}건 &nbsp;·&nbsp; 총 <b style="color:#555">{fmt_won(total,short=True)}</b></span>'
        f'</div>'
        f'{rows}</div>'
    )

_cc1, _cc2, _cc3 = st.columns(3, gap="medium")
with _cc1:
    st.markdown(_cmp_card(f"{mo_label(sel)} 신규 연체", _신규, "#e74c3c", "+", "#c0392b"), unsafe_allow_html=True)
with _cc2:
    st.markdown(_cmp_card(f"{mo_label(sel)} 추가 연체", _추가, "#e67e22", "+", "#c96a10"), unsafe_allow_html=True)
with _cc3:
    st.markdown(_cmp_card(f"{mo_label(sel)} 수금 완료", _수금, "#27ae60", "-", "#1a7a42"), unsafe_allow_html=True)

# ── 청구 발생 거래처 랭킹 ─────────────────────────────────────
with st.container(border=True):
    st.markdown(f'<div class="sec-hd"><span class="sec-bar"></span>{mo_label(sel)} 청구 발생 거래처</div>', unsafe_allow_html=True)
    charge = sel_df[sel_df["지체보상금"] > 0].sort_values("지체보상금", ascending=False).reset_index(drop=True)
    all_max = result_df[result_df["지체보상금"]>0]["지체보상금"].max() if not result_df[result_df["지체보상금"]>0].empty else 1

    if charge.empty:
        st.markdown('<div style="text-align:center;color:#ccc;padding:3rem 0;font-size:.9rem">해당 월 청구 발생 없음</div>', unsafe_allow_html=True)
    else:
        rank_colors = ["#c0392b","#e74c3c","#e67e22","#d4ac0d","#2980b9","#2471a3","#1a5276","#196f3d","#1e8449","#117a65"]
        rows_html = ""
        for i, row in charge.iterrows():
            v = int(row["지체보상금"])
            pct = min(v / all_max * 100, 100)
            color = rank_colors[i] if i < len(rank_colors) else "#888"
            사무소 = str(row.get("사무소","")).strip()
            사업부 = str(row.get("사업부","")).strip()
            담당자 = str(row.get("담당자","")).strip()
            # 표시할 조직 정보: 사무소 > 사업부 순으로 첫 번째 유효값
            org = 사무소 if 사무소 and 사무소 not in ("nan","") else (사업부 if 사업부 and 사업부 not in ("nan","") else "")
            org_badge = (
                f'<span style="background:#f0f2f5;color:#666;font-size:.7rem;'
                f'padding:.1rem .45rem;border-radius:3px;margin-left:.4rem">{org}</span>'
            ) if org else ""
            mgr_txt = (
                f'<span style="font-size:.72rem;color:#aaa;margin-left:.3rem">· {담당자}</span>'
            ) if 담당자 and 담당자 not in ("nan","") else ""
            rank_badge = (
                f'<span style="background:{color};color:#fff;font-size:.72rem;font-weight:800;'
                f'width:1.4rem;height:1.4rem;border-radius:50%;display:inline-flex;'
                f'align-items:center;justify-content:center;flex-shrink:0">{i+1}</span>'
            )
            rows_html += f"""
            <div style="display:grid;grid-template-columns:2rem 1fr 5rem;align-items:center;
                        gap:.8rem;padding:.55rem 0;border-bottom:1px solid #f5f5f5">
              {rank_badge}
              <div>
                <div style="font-size:.85rem;font-weight:600;color:#1a1a1a;margin-bottom:.25rem">
                  {row['판매처명']}{org_badge}{mgr_txt}
                </div>
                <div style="background:#f0f2f5;border-radius:4px;height:7px;overflow:hidden">
                  <div style="background:{color};width:{pct:.1f}%;height:100%;border-radius:4px"></div>
                </div>
              </div>
              <div style="text-align:right;font-size:.88rem;font-weight:700;color:{color}">{fmt_won(v,short=True)}</div>
            </div>"""
        st.markdown(f'<div style="padding:.2rem 0">{rows_html}</div>', unsafe_allow_html=True)

# ── 월별 + 누적 콤보 차트 (전체 너비) ────────────────────────
sel_yr_str = st.session_state.get("sel_yr", sel.split("-")[0])
all_12 = [f"{sel_yr_str}-{mo:02d}" for mo in range(1, 13)]
x_labels = [f"{int(m.split('-')[1])}월" for m in all_12]
bar_vals = [int(mo_sum.get(m, 0)) / 10000 for m in all_12]
cum_vals = []
run = 0
for v in bar_vals:
    run += v
    cum_vals.append(round(run, 1))
bar_colors = ["#e8604c" if m == sel else "#f2a99a" for m in all_12]
all_mo_max = max(bar_vals) if max(bar_vals) > 0 else 1
cum_max = max(cum_vals) if max(cum_vals) > 0 else 1
last_data_idx = max((i for i, v in enumerate(bar_vals) if v > 0), default=-1)
cum_line = [cum_vals[i] if i <= last_data_idx else None for i in range(12)]
cum_pts  = [cum_vals[i] if i <= last_data_idx and cum_vals[i] > 0 else None for i in range(12)]

with st.container(border=True):
    st.markdown(
        f'<div style="font-weight:800;font-size:.9rem;color:#1a2535;margin-bottom:.05rem">'
        f'{sel_yr_str}년 월별 · 누적 지체보상금</div>'
        f'<div style="font-size:.7rem;color:#aaa;margin-bottom:.5rem">단위: 만원 &nbsp;|&nbsp; '
        f'<span style="color:#f2a99a">■</span> 월별 발생 &nbsp;'
        f'<span style="color:#e67e22">─●</span> 누적</div>',
        unsafe_allow_html=True)
    fig_combo = go.Figure()
    # 누적 면적 (뒤에 깔리게 먼저)
    fig_combo.add_trace(go.Scatter(
        x=x_labels, y=cum_line, name="_fill",
        mode="lines", line=dict(color="rgba(0,0,0,0)", width=0),
        fill="tozeroy", fillcolor="rgba(232,146,74,0.10)",
        hoverinfo="skip", yaxis="y2", showlegend=False,
    ))
    # 바: 월별
    fig_combo.add_trace(go.Bar(
        x=x_labels, y=bar_vals, name="월별 발생",
        marker_color=bar_colors, marker_line_width=0,
        text=[f"{v:,.0f}만" if v > 0 else "" for v in bar_vals],
        textposition="outside", cliponaxis=False,
        textfont=dict(size=10, color="#888"),
        hovertemplate="<b>%{x}</b> 월별 %{y:,.0f}만원<extra></extra>",
        yaxis="y1",
    ))
    # 누적 라인+점+라벨
    fig_combo.add_trace(go.Scatter(
        x=x_labels, y=cum_line, name="누적 지체보상금",
        mode="lines+markers+text",
        line=dict(color="#e8924a", width=2.5),
        marker=dict(size=7, color="#e8924a", line=dict(color="#fff", width=2)),
        text=[f"{v:,.0f}만" if v else "" for v in cum_pts],
        textposition="top center",
        textfont=dict(size=10, color="#c07030"),
        hovertemplate="<b>%{x}</b> 누적 %{y:,.0f}만원<extra></extra>",
        yaxis="y2",
    ))
    fig_combo.update_layout(
        plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(t=45, b=10, l=55, r=65),
        height=350,
        bargap=0.4,
        legend=dict(
            orientation="h", x=0.5, xanchor="center", y=1.07,
            font=dict(size=11, family=FONT),
            bgcolor="rgba(0,0,0,0)",
        ),
        yaxis=dict(gridcolor="#f0f2f6", tickformat=",d", zeroline=False,
                   range=[0, all_mo_max * 1.55], tickfont=dict(size=10),
                   title=None, showline=False),
        yaxis2=dict(overlaying="y", side="right", range=[0, cum_max * 1.3],
                    tickformat=",d", tickfont=dict(size=10),
                    showgrid=False, title=None, showline=False),
        xaxis=dict(showgrid=False, tickfont=dict(size=11), showline=False, zeroline=False),
        font=dict(family=FONT),
    )
    st.plotly_chart(fig_combo, use_container_width=True, config={"displayModeBar": False})

# ── 다운로드 버튼 ────────────────────────────────────────────
_, dc2, dc3, _ = st.columns([1, 2, 2, 1])
with dc2:
    charge_only = sel_df[sel_df["지체보상금"] > 0].sort_values("지체보상금", ascending=False)
    st.download_button(
        f"⬇ {mo_label(sel)} 발생현황 엑셀",
        data=make_excel(charge_only, f"{mo_label(sel)}"),
        file_name=f"지체보상금_{sel.replace('-','')}_발생현황.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with dc3:
    st.download_button(
        "⬇ 전체 기간 엑셀",
        data=make_full_excel(result_df),
        file_name=f"지체보상금_전체_{date.today():%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── 상세 탭 ──────────────────────────────────────────────────
with st.container(border=True):
    tab1, tab2 = st.tabs([f"📋 {mo_label(sel)} 상세 내역", "📊 거래처별 누계 (전 기간)"])

    with tab1:
        fc1, fc2 = st.columns([4, 1])
        kw     = fc1.text_input("거래처명 검색", placeholder="거래처명 검색",
                                label_visibility="collapsed", key="kw1")
        only_c = fc2.checkbox("청구만", key="oc1")
        view   = sel_df.copy()
        if kw:     view = view[view["판매처명"].str.contains(kw, na=False)]
        if only_c: view = view[view["지체보상금"] > 0]
        view = view.sort_values("지체보상금", ascending=False).reset_index(drop=True)

        def _calc_str(r):
            if r["청구 증분일수"] <= 0: return "—"
            return (f'{int(r["현 미수금"]):,} × {int(r["청구 증분일수"])}일 × {r["요율"]}'
                    f' = {int(r["지체보상금"]):,}원')
        view["계산 근거"] = view.apply(_calc_str, axis=1)

        disp = view[["판매처명","현 미수금","유효기준회전일","현 회전일",
                     "지연일수(누적)","청구 증분일수","요율","지체보상금","계산 근거","비고"]]

        def _sty(r):
            if r["지체보상금"] > 0: return ["background:#fff8f0;font-weight:bold"] * len(r)
            return ["color:#ccc"] * len(r)

        disp = disp.reset_index(drop=True)
        disp.index = disp.index + 1
        st.dataframe(
            disp.style.apply(_sty, axis=1)
                .format({"현 미수금": "{:,.0f}", "지체보상금": "{:,.0f}"}),
            use_container_width=True, height=430,
        )

    with tab2:
        pv = result_df.pivot_table(
            index=["판매처코드","판매처명"], columns="기준월",
            values="지체보상금", aggfunc="sum", fill_value=0
        ).reset_index()
        pv.columns = [mo_label(c) if c not in ["판매처코드","판매처명"] else c for c in pv.columns]
        pv["합계"] = pv.iloc[:, 2:].sum(axis=1)
        pv = pv.sort_values("합계", ascending=False).reset_index(drop=True)
        pv.index = pv.index + 1
        nc = [c for c in pv.columns if c not in ["판매처코드","판매처명"]]
        def _heat(s):
            mx = s.max() if s.max() else 1
            return [f"background-color: rgba(255,100,50,{v/mx*0.6:.2f})" for v in s]
        st.dataframe(
            pv.style.format({c: "{:,.0f}" for c in nc})
              .apply(_heat, subset=["합계"]),
            use_container_width=True, height=430,
        )
