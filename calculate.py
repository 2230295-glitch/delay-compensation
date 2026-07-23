"""
지체보상금 자동 계산 (26년도 DW11 씽크채널)
- input/월별잔고회전일/ : 씽크채널 월말잔고_26년도 월별.xlsx (년월 컬럼 포함 단일 파일)
- input/기준회전일/     : SAP 거래처별 기준회전일_DW11_0722.xlsx
- output/              : 계산 결과 저장
"""

import sys
from pathlib import Path
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(__file__).parent
IN_M = BASE / "input" / "월별잔고회전일"
IN_B = BASE / "input" / "기준회전일"
OUT  = BASE / "output"

# 데이터 가공 기준 필터 설정
ALLOW_CH       = {"DW04","DW07","DW08","DW09","DW10","DW11","DW17"}
EXCLUDE_CODES  = {"10014461","C4900","C2000","C3100"}
EXCLUDE_BU     = {"기타사업부","수탁사업부","완제사업부"}
EXCLUDE_OFFICE = {"DW_B2B","DW_CH신사업팀"}
EXCLUDE_MGR    = {"1202150261"}


def _num(val) -> float:
    try:
        return float(str(val).replace(",","").strip())
    except Exception:
        return float("nan")


# ────────────────────────────────────────────
# 기준회전일 파일 로드
# ────────────────────────────────────────────
def load_base(folder: Path) -> pd.DataFrame:
    files = sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xls"))
    if not files:
        raise FileNotFoundError(f"기준회전일 파일이 없습니다: {folder}")

    dfs = [pd.read_excel(f, dtype=str).dropna(how="all") for f in files]
    df  = pd.concat(dfs, ignore_index=True)
    df.columns = [str(c).strip() for c in df.columns]

    out = []
    for _, r in df.iterrows():
        code = str(r.get("판매처", "")).strip()
        name = str(r.get("판매처명", "")).strip()
        base = _num(r.get("기준회전일", 0))
        yak  = _num(r.get("약정회전일", 0))
        exc  = _num(r.get("예외회전일", 0))

        if not code or code == "nan" or pd.isna(base) or base <= 0:
            continue

        # 유효 기준회전일: 예외 > 약정 > 기준 순서로 우선
        if not pd.isna(exc) and exc > 0:
            effective = int(exc)
        elif not pd.isna(yak) and yak > 0:
            effective = int(yak)
        else:
            effective = int(base)

        out.append({
            "코드":        code,
            "명칭":        name,
            "기준회전일":  int(base),
            "유효기준회전일": effective,
            "요율":        1 / 1000,
        })

    result = pd.DataFrame(out)
    print(f"  기준회전일: {len(result)}건 로드")
    return result


# ────────────────────────────────────────────
# 월별 잔고 파일 로드 (년월 컬럼 포함 단일 파일)
# ────────────────────────────────────────────
def load_monthly(folder: Path) -> list:
    # 기준/가공 파일 제외
    files = [
        f for f in sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xls"))
        if "기준" not in f.name and "가공" not in f.name
    ]
    if not files:
        raise FileNotFoundError(f"월별 잔고 파일이 없습니다: {folder}")

    dfs = []
    for f in files:
        tmp = pd.read_excel(f, dtype=str).dropna(how="all")
        tmp.columns = [str(c).strip() for c in tmp.columns]
        dfs.append(tmp)

    df = pd.concat(dfs, ignore_index=True)

    # 년월 파싱: "2026.01" → (2026, 1)
    if "년월" not in df.columns:
        raise ValueError(f"'년월' 컬럼 없음. 실제 컬럼: {list(df.columns)}")

    def _parse_ym(val):
        s = str(val).strip()
        if "." in s:
            parts = s.split(".")
            return int(parts[0]), int(parts[1])
        return None, None

    df["_yr"], df["_mo"] = zip(*df["년월"].map(_parse_ym))
    df = df[df["_yr"].notna() & df["_mo"].notna()].copy()
    df["_yr"] = df["_yr"].astype(int)
    df["_mo"] = df["_mo"].astype(int)

    # 잔고·회전일 숫자 변환
    df["_잔고"]   = df["순실잔고"].apply(_num)
    df["_회전일"] = df["순실회전일"].apply(_num)
    df = df.dropna(subset=["_잔고", "_회전일"]).copy()
    df["_회전일"] = df["_회전일"].astype(int)

    # ── 데이터 가공 기준 필터 적용 ──
    before = len(df)
    # ① 판매처 코드 없는 행
    df = df[df["판매처"].notna() & (df["판매처"].str.strip() != "") & (df["판매처"] != "nan")]
    # ② 채널 코드 필터
    df = df[df["채널"].isin(ALLOW_CH)]
    # ③ 순실잔고 = 0
    df = df[df["_잔고"] != 0]
    # ④ 담당자 코드 제외
    if "담당자" in df.columns:
        df = df[~df["담당자"].isin(EXCLUDE_MGR)]
    # ⑤ 사업부명 제외
    if "사업부명" in df.columns:
        df = df[~df["사업부명"].apply(
            lambda x: any(k in str(x) for k in EXCLUDE_BU)
        )]
    # ⑥ 특정 판매처 코드 제외
    df = df[~df["판매처"].isin(EXCLUDE_CODES)]
    # ⑦ 사무소명 제외
    if "사무소명" in df.columns:
        df = df[~df["사무소명"].isin(EXCLUDE_OFFICE)]

    print(f"  필터 전 {before}행 → 필터 후 {len(df)}행")

    # 월별로 분리
    months = []
    for (yr, mo), grp in df.groupby(["_yr", "_mo"]):
        rows = [{
            "연":     int(yr),
            "월":     int(mo),
            "코드":   str(r["판매처"]).strip(),
            "명칭":   str(r.get("판매처명", "")).strip(),
            "채널":   str(r.get("채널", "")).strip(),
            "잔고":   r["_잔고"],
            "현회전일": r["_회전일"],
        } for _, r in grp.iterrows()]
        months.append({"yr": int(yr), "mo": int(mo), "data": pd.DataFrame(rows)})
        print(f"  {yr}년 {mo:02d}월: {len(rows)}건")

    months.sort(key=lambda x: (x["yr"], x["mo"]))
    return months


# ────────────────────────────────────────────
# 지체보상금 계산
# ────────────────────────────────────────────
def calculate(months: list, base_df: pd.DataFrame) -> pd.DataFrame:
    base_map = {r["코드"]: r for _, r in base_df.iterrows()}

    # 거래처별 누적 최대 청구 지체일수
    cumulative: dict = {}  # 판매처코드 → max charged delay days
    results = []

    for m in months:
        yr, mo = m["yr"], m["mo"]
        for _, row in m["data"].iterrows():
            code = row["코드"]

            if code not in base_map:
                continue  # 기준회전일 미등록 → 스킵

            b        = base_map[code]
            기준일   = b["유효기준회전일"]
            요율     = b["요율"]
            현회전일  = row["현회전일"]
            잔고      = row["잔고"]

            지연일수  = max(0, 현회전일 - 기준일)
            prev_max  = cumulative.get(code, 0)

            if 지연일수 <= 0:
                cumulative[code] = 0
                증분일수 = 0
            else:
                증분일수 = max(0, 지연일수 - prev_max)
                cumulative[code] = max(prev_max, 지연일수)

            지체보상금 = round(잔고 * 증분일수 * 요율) if 증분일수 > 0 else 0

            results.append({
                "기준월":          f"{yr}-{mo:02d}",
                "판매처코드":      code,
                "판매처명":        row["명칭"],
                "채널":            row["채널"],
                "현 미수금":       int(잔고),
                "기준회전일(SAP)": b["기준회전일"],
                "유효기준회전일":  기준일,
                "현 회전일":       현회전일,
                "지연일수(누적)":  지연일수,
                "청구 증분일수":   증분일수,
                "지체보상금":      지체보상금,
                "요율":            f"1/{int(1/요율)}",
                "비고":            "청구" if 지체보상금 > 0 else (
                                   "정상화" if 지연일수 == 0 and prev_max > 0 else ""
                                   ),
            })

    return pd.DataFrame(results)


# ────────────────────────────────────────────
# 엑셀 출력
# ────────────────────────────────────────────
def export(df: pd.DataFrame, out_dir: Path):
    if df.empty:
        print("계산 결과 없음")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    path  = out_dir / f"지체보상금_계산결과_{today}.xlsx"

    wb  = openpyxl.Workbook()
    hdr = PatternFill("solid", fgColor="2E4A6E")

    def _header(ws):
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hdr
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 시트1: 전체 내역 ──
    ws1   = wb.active
    ws1.title = "전체내역"
    cols  = list(df.columns)
    ws1.append(cols)
    for r in df.itertuples(index=False):
        ws1.append(list(r))
    _header(ws1)

    money = {"현 미수금", "지체보상금"}
    for ci, cn in enumerate(cols, 1):
        for ri in range(2, ws1.max_row + 1):
            cell = ws1.cell(ri, ci)
            if cn in money:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")

    widths = {
        "기준월": 10, "판매처코드": 12, "판매처명": 22, "채널": 10,
        "현 미수금": 14, "기준회전일(SAP)": 12, "유효기준회전일": 12,
        "현 회전일": 10, "지연일수(누적)": 12, "청구 증분일수": 12,
        "지체보상금": 14, "요율": 8, "비고": 8,
    }
    for i, c in enumerate(cols, 1):
        ws1.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)
    ws1.freeze_panes = "A2"

    # ── 시트2: 거래처별 월별 요약 ──
    ws2 = wb.create_sheet("거래처별요약")
    pivot = df.pivot_table(
        index=["판매처코드", "판매처명"],
        columns="기준월",
        values="지체보상금",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pcols = list(pivot.columns)
    ws2.append(pcols)
    for r in pivot.itertuples(index=False):
        ws2.append(list(r))
    _header(ws2)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    for i in range(3, len(pcols) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = "C2"

    # ── 시트3: 월별 합계 ──
    ws3 = wb.create_sheet("월별합계")
    monthly = df.groupby("기준월").agg(
        청구건수      =("지체보상금", lambda x: (x > 0).sum()),
        지체보상금합계=("지체보상금", "sum"),
        청구거래처수  =("판매처코드", lambda x: x[df.loc[x.index, "지체보상금"] > 0].nunique()),
    ).reset_index()
    ws3.append(list(monthly.columns))
    for r in monthly.itertuples(index=False):
        ws3.append(list(r))
    _header(ws3)
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")
    for i in range(1, 5):
        ws3.column_dimensions[get_column_letter(i)].width = 18
    ws3.freeze_panes = "A2"

    wb.save(path)
    print(f"\n결과 저장: {path}")
    print(f"  전체내역: {len(df)}행")
    print(f"  청구 발생: {(df['지체보상금'] > 0).sum()}건")
    print(f"  총 지체보상금: {df['지체보상금'].sum():,.0f}원")


# ────────────────────────────────────────────
# 실행
# ────────────────────────────────────────────
def main():
    print("=== 지체보상금 계산 시작 ===\n")

    print("[1] 기준회전일 파일 로드")
    base_df = load_base(IN_B)

    print("\n[2] 월별 잔고/회전일 파일 로드")
    months = load_monthly(IN_M)
    if not months:
        print("월별 파일 없음 — 종료")
        return

    print("\n[3] 지체보상금 계산")
    result = calculate(months, base_df)

    print("\n[4] 결과 출력")
    export(result, OUT)


if __name__ == "__main__":
    main()
